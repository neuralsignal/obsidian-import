"""XLSX text extraction using openpyxl.

Extracts sheet-by-sheet data as markdown tables.
"""

from __future__ import annotations

from pathlib import Path

from obsidian_import.formatting import render_markdown_table
from obsidian_import.timeout import run_with_timeout


def extract(path: Path, timeout_seconds: int, max_rows_per_sheet: int) -> str:
    """Extract data from an XLSX file, returning markdown."""
    return run_with_timeout(lambda: _extract_xlsx(path, max_rows_per_sheet), timeout_seconds, "XLSX", path)


def _extract_xlsx(path: Path, max_rows_per_sheet: int) -> str:
    """Internal XLSX extraction logic."""
    from openpyxl import load_workbook

    wb = load_workbook(str(path), data_only=True, read_only=True)
    sections: list[str] = [f"# {path.stem}"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[list[str]] = []
        truncated = False

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx >= max_rows_per_sheet:
                truncated = True
                break
            cells = [str(cell) if cell is not None else "" for cell in row]
            if any(c.strip() for c in cells):
                rows.append(cells)

        if not rows:
            continue

        sheet_sections: list[str] = [f"## Sheet: {sheet_name}"]

        sheet_sections.append(render_markdown_table(rows))

        if truncated:
            sheet_sections.append(f"\n*Truncated: sheet has more than {max_rows_per_sheet} rows.*")

        sections.append("\n\n".join(sheet_sections))

    wb.close()
    return "\n\n".join(sections)
